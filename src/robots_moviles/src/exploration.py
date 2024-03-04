#!/usr/bin/env python
import rospy
import actionlib
from move_base_msgs.msg import MoveBaseAction, MoveBaseGoal
from nav_msgs.msg import OccupancyGrid
from random import randint

import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import numpy as np
import math
import openpyxl

# Definir un tamaño para el área cuadrada alrededor del punto más lejano desconocido
SEARCH_AREA_SIZE = 10  # Tamaño del área cuadrada: 3x3 MINIMO
# Definir los valores de los obstáculos y del área desconocida
OBSTACLE_VALUE = 100
UNKNOWN_VALUE = -1
CHANGE_VALUE = 50
map_data = None
counter = 0
origin_x = 0
origin_y = 0

def map_callback(map_msg):
    global map_data, counter
    counter += 1
    map_data = map_msg
    print(f'Mapa recibido {counter}')
    map_matriz = np.array(map_data.data).reshape((map_data.info.height, map_data.info.width))
    area_mapeada(map_matriz)
    #save_map_to_excel('map_data_' + str(counter) + '.xlsx')
    #save_map_imagen('map_data_' + str(counter) + '.xlsx')    
    select_and_publish_goal()

#FUNCION PARA GUARDAR EN ARCHIVO .xlsx
def save_map_to_excel(filename):
    if map_data is not None:
        wb = openpyxl.Workbook()
        ws = wb.active
        width = map_data.info.width
        height = map_data.info.height
        for y in range(height):
            for x in range(width):
                index = y * width + x
                ws.cell(row=y+1, column=x+1, value=map_data.data[index])
        wb.save(filename)
        print(f'Mapa guardado en {filename}')

#FUNCION PARA GUARDAR EN ARCHIVO .png
def save_map_imagen(file_path):
    # Cargar datos desde el archivo Excel
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    # Leer datos de la hoja activa
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    # Convertir los datos a una matriz numpy
    data = np.array(data)
    # Mapear los valores a colores
    cmap_colors = [(0.5, 0.5, 0.5), (0, 0, 1), (1, 1, 1), (1, 0, 0)]  # Gris, Azul, Blanco, Rojo
    cmap = ListedColormap(cmap_colors)
    normalized_map_data = data.astype(float) / 100  # Normalizamos al rango [0, 1]
    # Graficar los datos
    plt.imshow(normalized_map_data, cmap=cmap, origin='lower', vmin=-0.5, vmax=1.5)  # Ajustar los límites de los valores para que los colores coincidan
    #plt.colorbar()
    plt.savefig('map_data_' + str(counter) + '.png')  # Guardar la imagen como un archivo PNG
    print(f'Imagen guardada map_data_' + str(counter) + '.png')
    plt.close()

# FUNCION PARA EXTRAER LA REGION DESEADA A EXPLORAR
def area_mapeada(map_matriz):
    global map_data  # Declarar map_data como global si es necesario
    if map_matriz is not None:
        width = map_matriz.shape[1]
        height = map_matriz.shape[0]
        border_indices = np.where(map_matriz == 100)
        min_row, max_row = min(border_indices[0]), max(border_indices[0])
        min_col, max_col = min(border_indices[1]), max(border_indices[1])       
        # Convertir map_data.data en una lista para poder modificarla
        map_data_list = list(map_data.data)
        # Asignar el valor 50 a las celdas fuera del área reducida que no tienen el valor 100
        for y in range(height):
            for x in range(width):
                index = y * width + x
                if x < min_col or x > max_col or y < min_row or y > max_row:
                    if map_matriz[y][x] != OBSTACLE_VALUE:
                        map_data_list[index] = CHANGE_VALUE      
        # Convertir map_data_list nuevamente en una tupla
        map_data.data = tuple(map_data_list)
        print('Mapa modificado')

#FUNCION DE LA ESTRATEGIA
def select_and_publish_goal():
    global origin_x, origin_y
    max_distance = 0
    target_x, target_y = None, None
    map_matriz = np.array(map_data.data).reshape((map_data.info.height, map_data.info.width))
    if counter == 0:
        origin_x = abs(int(map_data.info.origin.position.x/0.05))
        origin_y = abs(int(map_data.info.origin.position.y/0.05))
    if map_data is not None:
        # Recorrer todas las celdas del mapa
        for y in range(map_data.info.height):
            for x in range(map_data.info.width):
                index = int(y * map_data.info.width + x)
                if map_data.data[index] == UNKNOWN_VALUE:
                    # Calcular la distancia euclidiana desde la posición actual del robot
                    distance = math.sqrt((origin_x - x)**2 + (origin_y - y)**2)
                    if distance > max_distance:
                        # Verificar el área cuadrada alrededor de la celda actual
                        if check_neighborhood(map_data, x, y):
                            max_distance = distance
                            target_x, target_y = x, y
        if target_x is not None and target_y is not None:
            # Seleccionar el punto más lejano desconocido que cumpla con las condiciones
            print(f'Punto de partida: ({origin_y}, {origin_x})') 
            send_goal(map_data, target_x, target_y)
            origin_x = target_x
            origin_y = target_y
            #print(f'Porcentaje de celdas desconocidas: {unknown_percentage:.2f}%')
        else:    
            print('No se encontraron áreas desconocidas en el mapa. Exploración completada.')
            rospy.signal_shutdown("Exploración completada.")

#FUNCION PARA MOVER EL ROBOT A LA POSICION DESEADA
def send_goal(map_data, target_x, target_y):           
    goal_client = actionlib.SimpleActionClient('move_base', MoveBaseAction)
    goal_client.wait_for_server()        
    goal = MoveBaseGoal()
    goal.target_pose.header.frame_id = "map"
    goal.target_pose.header.stamp = rospy.Time.now()
    go_x = target_x * map_data.info.resolution + map_data.info.origin.position.x
    go_y= target_y * map_data.info.resolution + map_data.info.origin.position.y
    print(f'Destino seleccionado: ({target_y}, {target_x}) Navegando hacia el punto...{go_y},{go_x}')  
    goal.target_pose.pose.position.x = go_x
    goal.target_pose.pose.position.y = go_y 
    goal.target_pose.pose.orientation.w = 1.0
    goal_client.send_goal(goal)
    wait = goal_client.wait_for_result()
    print(f'Punto alcanzado!')

#FUNCION PARA VERIFICAR SI NO EXISTEN OBSTACULOS EN EL PUNTO SELECCIONADO
def check_neighborhood(map_data, x, y):
    area_size = SEARCH_AREA_SIZE
    # Definir los límites del área cuadrada alrededor de la celda (x, y)
    x_min = max(0, x - area_size)
    x_max = min(map_data.info.width - 1, x + area_size)
    y_min = max(0, y - area_size)
    y_max = min(map_data.info.height - 1, y + area_size)
    # Verificar si todas las celdas dentro del área cuadrada son transitables
    for i in range(x_min, x_max + 1):
        for j in range(y_min, y_max + 1):
            index = j * map_data.info.width + i
            if map_data.data[index] == OBSTACLE_VALUE or map_data.data[index] == CHANGE_VALUE:
                return False
    return True

if __name__ == '__main__':   
    try:
        rospy.init_node('exploration', anonymous=True)
        map_subscriber = rospy.Subscriber('/map', OccupancyGrid, map_callback, queue_size=1)
        rate = rospy.Rate(10) 
        while not rospy.is_shutdown():          
            rate.sleep()
    except rospy.ROSInterruptException:
        save_map_to_excel('map_data_' + str(counter) + '.xlsx')
        save_map_imagen('map_data_' + str(counter) + '.xlsx')
        rospy.loginfo("Exploration finished.")
