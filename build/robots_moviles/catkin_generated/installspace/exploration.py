#!/usr/bin/env python3
<<<<<<< HEAD
=======

>>>>>>> 022097bbb771e3bb79f85a5fd84ee629f95f3c0b
import rospy
import actionlib
from move_base_msgs.msg import MoveBaseAction, MoveBaseGoal
from nav_msgs.msg import OccupancyGrid
from random import randint

<<<<<<< HEAD
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import numpy as np
import math
import openpyxl

def map_callback(map_msg):
    global counter
    print('Mapa recibido')
    #plot_map(map_msg,'map_img_' + str(counter) + '.png') 
    #save_map_data_to_excel(map_msg,'map_data_' + str(counter) + '.xlsx')
    #visualize_map_data('map_data_' + str(counter) + '.xlsx')
    map_matriz = np.array(map_msg.data).reshape((map_msg.info.height, map_msg.info.width))
    map_data2=extract_reduced_map(map_matriz,map_msg)
    #plot_map(map_data2)    
    counter += 1
    select_and_publish_goal(map_data2)

def save_map_data_to_excel(map_data, filename):
    if map_data is not None:
        wb = openpyxl.Workbook()
        ws = wb.active
        width = map_data.info.width
        height = map_data.info.height

        for y in range(height):
            for x in range(width):
                index = y * width + x
                ws.cell(row=y+1, column=x+1, value=map_data.data[index])
        wb.save(filename)
        print(f'Mapa guardado en {filename}')

def visualize_map_data(file_path):
    global counter
    # Cargar datos desde el archivo Excel
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    # Leer datos de la hoja activa
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    # Convertir los datos a una matriz numpy
    data = np.array(data)
    # Mapear los valores a colores
    cmap_colors = [(0.5, 0.5, 0.5), (0, 0, 1), (1, 1, 1), (1, 0, 0)]  # Gris, Azul, Blanco, Rojo
    cmap = ListedColormap(cmap_colors)
    normalized_map_data = data.astype(float) / 100  # Normalizamos al rango [0, 1]
    # Graficar los datos
    plt.imshow(normalized_map_data, cmap=cmap, origin='lower', vmin=-0.5, vmax=1.5)  # Ajustar los límites de los valores para que los colores coincidan
    #plt.colorbar()
    plt.savefig('map_data_' + str(counter) + '.png')  # Guardar la imagen como un archivo PNG
    print(f'Imagen guardada map_data_' + str(counter) + '.png')
    plt.close()

def plot_map(map_data,filename):
    if map_data is not None:
        width = map_data.info.width
        height = map_data.info.height
        data = map_data.data
        map_array = np.array(data).reshape((height, width))
        # Definir el mapa de colores personalizado
        cmap_colors = [(0.5, 0.5, 0.5), (0, 0, 1), (1, 1, 1), (1, 0, 0)]  # Gris, Azul, Blanco, Rojo
        cmap = ListedColormap(cmap_colors)
        # Normalizar el mapa de datos
        normalized_map_data = map_array.astype(float) / 100  # Normalizamos al rango [0, 1]
        # Graficar la matriz usando el mapa de colores personalizado
        plt.imshow(normalized_map_data, cmap=cmap, origin='lower', vmin=-0.5, vmax=1.5)  # Ajustar los límites de los valores para que los colores coincidan
        #plt.colorbar(label='Occupancy Status')
        plt.title('Occupancy Grid Map')
        plt.xlabel('X')
        plt.ylabel('Y')
        plt.savefig(filename)  # Guardar la imagen como un archivo PNG
        plt.close()
        print(f'Imagen guardada {filename}')

# Función para extraer la región reducida del mapa
def extract_reduced_map(map_matriz,map_data):
    if map_matriz is not None:
        print('Extrayendo región reducida...')
        width = map_matriz.shape[1]
        height = map_matriz.shape[0]
        border_indices = np.where(map_matriz == 100)
        min_row, max_row = min(border_indices[0]), max(border_indices[0])
        min_col, max_col = min(border_indices[1]), max(border_indices[1])
        # Asignar el valor 50 a las celdas fuera del área reducida que no tienen el valor 100
        map_matriz[:min_row][map_matriz[:min_row] != 100] = 50
        map_matriz[max_row+1:][map_matriz[max_row+1:] != 100] = 50
        map_matriz[:, :min_col][map_matriz[:, :min_col] != 100] = 50
        map_matriz[:, max_col+1:][map_matriz[:, max_col+1:] != 100] = 50
        map_data.data = list(map_data.data)
        for y in range(height):
            for x in range(width):
                index = int(y * width + x) 
                map_data.data[index]=map_matriz[y][x]
        return map_data
    
def select_and_publish_goal(map_msg):
    global map_data
    map_data = map_msg
=======
def map_callback(map_msg):
    print('Mapa recibido')
    map_data = map_msg
    select_and_publish_goal(map_data)

def select_and_publish_goal(map_data):
>>>>>>> 022097bbb771e3bb79f85a5fd84ee629f95f3c0b
    if map_data is not None:
        print('Eligiendo destino...')
        width = map_data.info.width
        height = map_data.info.height
<<<<<<< HEAD
        total_cells = width * height
        max_unknown_cells = 0
        best_x, best_y = 0, 0

        for y in range(height):
            for x in range(width):
                if map_data.data[y * width + x] == -1:
                    unknown_cells_count = count_unknown_cells(map_data, x, y)
                    if unknown_cells_count > max_unknown_cells:
                        max_unknown_cells = unknown_cells_count
                        best_x, best_y = x, y
        unknown_percentage = (unknown_cells_count / total_cells) * 100

        if max_unknown_cells > 0:
            print(f'Coordenada desconocida: ({best_x}, {best_y})')
            print('Buscando espacio libre más cercano...')
            closest_x, closest_y = find_closest_free_space(map_data, best_x, best_y)
            if closest_x is not None and closest_y is not None:
                print(f'Destino seleccionado: ({closest_x}, {closest_y}). Navegando hacia el punto...')            
                goal_client = actionlib.SimpleActionClient('move_base', MoveBaseAction)
                goal_client.wait_for_server()        
                goal = MoveBaseGoal()
                goal.target_pose.header.frame_id = "map"
                goal.target_pose.header.stamp = rospy.Time.now()
                goal.target_pose.pose.position.x = closest_x * map_data.info.resolution + map_data.info.origin.position.x
                goal.target_pose.pose.position.y = closest_y  * map_data.info.resolution + map_data.info.origin.position.y
                goal.target_pose.pose.orientation.w = 1.0
                goal_client.send_goal(goal)
                wait = goal_client.wait_for_result()
                print('Punto alcanzado!')
                print(f'Porcentaje de celdas desconocidas: {unknown_percentage:.2f}%')
            else:
                print('No se encontraron espacios libres cercanos al área con mayor concentración de celdas desconocidas.')
                rospy.signal_shutdown("Exploración completada.")
        else:
            print('No se encontraron áreas desconocidas en el mapa. Exploración completada.')
            rospy.signal_shutdown("Exploración completada.")

def count_unknown_cells(map_data, x, y):
    width = map_data.info.width
    height = map_data.info.height
    unknown_cells = 0

    # Definir los límites de la vecindad (por ejemplo, un cuadrado de 3x3 alrededor de la celda)
    neighborhood_size = 3
    min_x = max(0, x - neighborhood_size // 2)
    max_x = min(width - 1, x + neighborhood_size // 2)
    min_y = max(0, y - neighborhood_size // 2)
    max_y = min(height - 1, y + neighborhood_size // 2)

    for ny in range(min_y, max_y + 1):
        for nx in range(min_x, max_x + 1):
            if map_data.data[ny * width + nx] == -1:
                unknown_cells += 1

    return unknown_cells

def find_closest_free_space(map_data, x, y):
    width = map_data.info.width
    height = map_data.info.height
    min_distance = float('inf')
    closest_x, closest_y = None, None

    for dy in range(-1, 2):
        for dx in range(-1, 2):
            nx, ny = x + dx, y + dy
            if 0 <= nx < width and 0 <= ny < height and map_data.data[ny * width + nx] == 0:
                distance = math.sqrt((nx - x) ** 2 + (ny - y) ** 2)
                if distance < min_distance:
                    min_distance = distance
                    closest_x, closest_y = nx, ny

    return closest_x, closest_y

if __name__ == '__main__':   
    map_data = None
    try:
        rospy.init_node('exploration', anonymous=True)
        map_subscriber = rospy.Subscriber('/map', OccupancyGrid, map_callback,queue_size=1)
        counter=0
        rate = rospy.Rate(10) 
        while not rospy.is_shutdown():          
            rate.sleep()

    except rospy.ROSInterruptException:
        save_map_data_to_excel(map_data,'map_data_' + str(counter) + '.xlsx')
        visualize_map_data('map_data_' + str(counter) + '.xlsx')
=======

        goalFound = False;
        while not goalFound:
            random_x = randint(0, width - 1)
            random_y = randint(0, height - 1)
            index = random_y * width + random_x

            if map_data.data[index] == 0:
                goalFound = True;

        print('Destino elegido. Navegando hasta el punto...')

        goal_client = actionlib.SimpleActionClient('move_base',MoveBaseAction)
        goal_client.wait_for_server()
        
        goal = MoveBaseGoal()
        goal.target_pose.header.frame_id = "map"
        goal.target_pose.header.stamp = rospy.Time.now()
        goal.target_pose.pose.position.x = random_x * map_data.info.resolution + map_data.info.origin.position.x
        goal.target_pose.pose.position.y = random_y * map_data.info.resolution + map_data.info.origin.position.y
        goal.target_pose.pose.orientation.w = 1.0

        goal_client.send_goal(goal)
        wait = goal_client.wait_for_result()
        print('Punto alcanzado!')


if __name__ == '__main__':
    try:
        rospy.init_node('exploration', anonymous=True)
        map_data = None
        map_subscriber = rospy.Subscriber('/map', OccupancyGrid, map_callback)
        
        rate = rospy.Rate(1)

        while not rospy.is_shutdown():
            rate.sleep()

    except rospy.ROSInterruptException:
>>>>>>> 022097bbb771e3bb79f85a5fd84ee629f95f3c0b
        rospy.loginfo("Exploration finished.")
